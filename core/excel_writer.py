# report_maker/core/excel_writer.py
import io
from datetime import datetime
from typing import Dict, Optional
from openpyxl import load_workbook
from .settings import SHEET_NAME, JST
from .textutil import split_lines, sanitize_filename
from .parsing import try_parse_datetime, split_dt_components, first_date_yyyymmdd

def _fill_multiline(ws, col_letter: str, start_row: int, text: Optional[str], max_lines: int = 5):
    for i in range(max_lines):
        ws[f"{col_letter}{start_row + i}"] = ""
    if not text:
        return
    for idx, line in enumerate(split_lines(text, max_lines=max_lines)[:max_lines]):
        ws[f"{col_letter}{start_row + idx}"] = line

def fill_template_xlsx(template_bytes: bytes, data: Dict[str, Optional[str]]) -> bytes:
    if not template_bytes:
        raise ValueError("テンプレートのバイト列が空です。")

    try:
        wb = load_workbook(io.BytesIO(template_bytes), keep_vba=True)
    except Exception as e:
        raise RuntimeError(f"テンプレートの読み込みに失敗しました（破損の可能性）: {e}") from e

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    if data.get("管理番号"): ws["C12"] = data["管理番号"]
    if data.get("メーカー"): ws["J12"] = data["メーカー"]
    if data.get("制御方式"): ws["M12"] = data["制御方式"]
    if data.get("通報者"): ws["C14"] = data["通報者"]
    if data.get("対応者"): ws["L37"] = data["対応者"]

    pa = (data.get("処理修理後") or "").strip()
    if pa:
        ws["C35"] = pa

    if data.get("所属"):
        ws["C37"] = data["所属"]

    now = datetime.now(JST)
    ws["B5"], ws["D5"], ws["F5"] = now.year, now.month, now.day

    def write_dt_block(base_row: int, src_key: str):
        dt = try_parse_datetime(data.get(src_key))
        y, m, d, wd, hh, mm = split_dt_components(dt)
        cellmap = {"Y": f"C{base_row}", "Mo": f"F{base_row}", "D": f"H{base_row}",
                   "W": f"J{base_row}", "H": f"M{base_row}", "Min": f"O{base_row}"}
        if y is not None: ws[cellmap["Y"]] = y
        if m is not None: ws[cellmap["Mo"]] = m
        if d is not None: ws[cellmap["D"]] = d
        if wd is not None: ws[cellmap["W"]] = wd
        if hh is not None: ws[cellmap["H"]] = f"{hh:02d}"
        if mm is not None: ws[cellmap["Min"]] = f"{mm:02d}"

    write_dt_block(13, "受信時刻")
    write_dt_block(19, "現着時刻")
    write_dt_block(36, "完了時刻")

    _fill_multiline(ws, "C", 15, data.get("受信内容"), max_lines=4)
    _fill_multiline(ws, "C", 20, data.get("現着状況"))
    _fill_multiline(ws, "C", 25, data.get("原因"))
    _fill_multiline(ws, "C", 30, data.get("処置内容"))

    out = io.BytesIO()
    try:
        wb.save(out)
    except Exception as e:
        raise RuntimeError(f"Excel保存時に失敗しました: {e}") from e

    return out.getvalue()

def build_filename(data: Dict[str, Optional[str]]) -> str:
    base_day = first_date_yyyymmdd(data.get("現着時刻"), data.get("完了時刻"), data.get("受信時刻"))
    manageno = sanitize_filename((data.get("管理番号") or "UNKNOWN").strip().replace("/", "_"))
    bname = sanitize_filename((data.get("物件名") or "").strip().replace("/", "_"))
    return (f"緊急出動報告書_{manageno}_{bname}_{base_day}.xlsm" if bname
            else f"緊急出動報告書_{manageno}_{base_day}.xlsm")
